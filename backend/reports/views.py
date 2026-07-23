from io import BytesIO
from datetime import date
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle, Spacer, HRFlowable

from rest_framework.views import APIView
from vehicles.models import Vehicle
from maintenance.models import MaintenanceRecord
from trips.models import Trip
from ml.views import compute_prediction_details


class VehiclePDFReportAPIView(APIView):
    def get(self, request):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=15,
            fontName='Helvetica'
        )
        cell_style = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            fontName='Helvetica'
        )
        cell_header_style = ParagraphStyle(
            'CellHeader',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.white,
            fontName='Helvetica-Bold'
        )

        elements = []
        elements.append(Paragraph("Fleet Vehicle Manifest Report", title_style))
        elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y - %H:%M UTC')}", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#CBD5E1'), spaceAfter=15))

        vehicles = Vehicle.objects.all()

        table_data = [
            [
                Paragraph("<b>Reg No.</b>", cell_header_style),
                Paragraph("<b>Vehicle Name</b>", cell_header_style),
                Paragraph("<b>Type</b>", cell_header_style),
                Paragraph("<b>Mileage</b>", cell_header_style),
                Paragraph("<b>Status</b>", cell_header_style),
                Paragraph("<b>Insurance Expiry</b>", cell_header_style),
            ]
        ]

        for v in vehicles:
            table_data.append([
                Paragraph(v.registration_number, cell_style),
                Paragraph(v.vehicle_name, cell_style),
                Paragraph(v.get_vehicle_type_display(), cell_style),
                Paragraph(f"{v.mileage:,} km", cell_style),
                Paragraph(v.get_status_display(), cell_style),
                Paragraph(str(v.insurance_expiry), cell_style),
            ])

        t = Table(table_data, colWidths=[90, 110, 60, 80, 80, 110])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))

        elements.append(t)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="vehicles_manifest_report.pdf"'
        return response


class VehicleExcelReportAPIView(APIView):
    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fleet Vehicles"

        headers = ["Registration", "Vehicle Name", "Type", "Manufacturer", "Model", "Year", "Mileage (km)", "Status", "Driver", "Insurance Expiry"]
        sheet.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        vehicles = Vehicle.objects.all()

        for v in vehicles:
            driver_name = v.assigned_driver.name if v.assigned_driver else "Unassigned"
            sheet.append([
                v.registration_number,
                v.vehicle_name,
                v.get_vehicle_type_display(),
                v.manufacturer,
                v.model,
                v.manufacturing_year,
                v.mileage,
                v.get_status_display(),
                driver_name,
                str(v.insurance_expiry),
            ])

        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="vehicles_manifest_report.xlsx"'
        workbook.save(response)
        return response


class AIPredictionPDFReportAPIView(APIView):
    def get(self, request):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'AITitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=4,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'AISubtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=12,
            fontName='Helvetica'
        )
        metric_title_style = ParagraphStyle('MetricTitle', fontSize=9, textColor=colors.HexColor('#475569'), fontName='Helvetica-Bold')
        metric_val_style = ParagraphStyle('MetricVal', fontSize=14, textColor=colors.HexColor('#0F172A'), fontName='Helvetica-Bold')
        
        cell_style = ParagraphStyle('AICell', fontSize=8, leading=10, fontName='Helvetica')
        cell_header_style = ParagraphStyle('AICellHeader', fontSize=8, leading=10, textColor=colors.white, fontName='Helvetica-Bold')
        
        crit_badge = ParagraphStyle('CritBadge', fontSize=8, leading=10, textColor=colors.HexColor('#DC2626'), fontName='Helvetica-Bold')
        high_badge = ParagraphStyle('HighBadge', fontSize=8, leading=10, textColor=colors.HexColor('#EA580C'), fontName='Helvetica-Bold')
        mod_badge = ParagraphStyle('ModBadge', fontSize=8, leading=10, textColor=colors.HexColor('#D97706'), fontName='Helvetica-Bold')
        opt_badge = ParagraphStyle('OptBadge', fontSize=8, leading=10, textColor=colors.HexColor('#16A34A'), fontName='Helvetica-Bold')

        elements = []
        elements.append(Paragraph("AI Predictive Maintenance & Fleet Risk Report", title_style))
        elements.append(Paragraph(f"Machine Learning Telemetry Audit | Generated: {timezone.now().strftime('%B %d, %Y - %H:%M UTC')}", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#6366F1'), spaceAfter=12))

        vehicles = Vehicle.objects.all()
        now = timezone.now().date()

        evaluated_data = []
        crit_count = 0
        high_count = 0
        opt_count = 0
        total_prob = 0.0

        for v in vehicles:
            mileage = float(v.mileage)
            year = int(v.manufacturing_year)
            service_count = MaintenanceRecord.objects.filter(vehicle=v).count()
            trips = Trip.objects.filter(vehicle=v, status="COMPLETED")
            total_trip_distance = sum(t.distance for t in trips)
            
            recent_completed = MaintenanceRecord.objects.filter(vehicle=v, status="COMPLETED").order_by("-service_date").first()
            days_since = (now - recent_completed.service_date).days if recent_completed else int((2026 - year) * 365 / 2)
            
            diag = compute_prediction_details(
                mileage=mileage,
                year=year,
                service_count=service_count,
                total_trip_distance=total_trip_distance,
                days_since_last_service=days_since
            )
            diag["vehicle"] = v
            evaluated_data.append(diag)

            prob = diag["failure_probability"]
            total_prob += prob
            if diag["risk_level"] == "Critical":
                crit_count += 1
            elif diag["risk_level"] == "High":
                high_count += 1
            else:
                opt_count += 1

        avg_risk = round(total_prob / max(1, len(vehicles)), 1)

        # Overview Metrics Box Table
        summary_table_data = [
            [
                Paragraph("Total Fleet Units", metric_title_style),
                Paragraph("Avg Failure Risk", metric_title_style),
                Paragraph("Critical Risk Units", metric_title_style),
                Paragraph("High Risk Units", metric_title_style),
                Paragraph("Optimal Units", metric_title_style),
            ],
            [
                Paragraph(str(len(vehicles)), metric_val_style),
                Paragraph(f"{avg_risk}%", metric_val_style),
                Paragraph(f"<font color='#DC2626'>{crit_count}</font>", metric_val_style),
                Paragraph(f"<font color='#EA580C'>{high_count}</font>", metric_val_style),
                Paragraph(f"<font color='#16A34A'>{opt_count}</font>", metric_val_style),
            ]
        ]
        sum_table = Table(summary_table_data, colWidths=[108, 108, 108, 108, 108])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F5F9')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))

        elements.append(sum_table)
        elements.append(Spacer(1, 15))

        # Fleet Risk Audit Table
        table_data = [
            [
                Paragraph("<b>Reg No.</b>", cell_header_style),
                Paragraph("<b>Vehicle Name</b>", cell_header_style),
                Paragraph("<b>Mileage</b>", cell_header_style),
                Paragraph("<b>Risk Prob.</b>", cell_header_style),
                Paragraph("<b>Risk Tier</b>", cell_header_style),
                Paragraph("<b>AI Actionable Recommendation</b>", cell_header_style),
            ]
        ]

        for item in evaluated_data:
            v = item["vehicle"]
            risk = item["risk_level"]
            if risk == "Critical":
                badge = crit_badge
            elif risk == "High":
                badge = high_badge
            elif risk == "Moderate":
                badge = mod_badge
            else:
                badge = opt_badge

            table_data.append([
                Paragraph(v.registration_number, cell_style),
                Paragraph(f"{v.vehicle_name}<br/><font color='#64748B'>{v.get_vehicle_type_display()}</font>", cell_style),
                Paragraph(f"{item['mileage']:,} km<br/><font color='#64748B'>{item['year']}</font>", cell_style),
                Paragraph(f"<b>{item['failure_probability']}%</b>", cell_style),
                Paragraph(item["risk_level"], badge),
                Paragraph(item["recommended_action"], cell_style),
            ])

        t = Table(table_data, colWidths=[75, 95, 70, 60, 60, 180])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#312E81')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))

        elements.append(t)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="ai_predictive_maintenance_report.pdf"'
        return response


class AIPredictionExcelReportAPIView(APIView):
    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "AI Maintenance Audit"

        headers = [
            "Registration Number",
            "Vehicle Name",
            "Type",
            "Mileage (km)",
            "Manufacturing Year",
            "Age (years)",
            "Completed Services Count",
            "Days Since Last Service",
            "Failure Probability (%)",
            "Risk Tier",
            "Prediction Verdict",
            "Recommended Action",
            "Est. Maintenance Cost (₹)"
        ]
        sheet.append(headers)

        header_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        vehicles = Vehicle.objects.all()
        now = timezone.now().date()

        for v in vehicles:
            mileage = float(v.mileage)
            year = int(v.manufacturing_year)
            service_count = MaintenanceRecord.objects.filter(vehicle=v).count()
            trips = Trip.objects.filter(vehicle=v, status="COMPLETED")
            total_trip_distance = sum(t.distance for t in trips)

            recent_completed = MaintenanceRecord.objects.filter(vehicle=v, status="COMPLETED").order_by("-service_date").first()
            days_since = (now - recent_completed.service_date).days if recent_completed else int((2026 - year) * 365 / 2)

            diag = compute_prediction_details(
                mileage=mileage,
                year=year,
                service_count=service_count,
                total_trip_distance=total_trip_distance,
                days_since_last_service=days_since
            )

            sheet.append([
                v.registration_number,
                v.vehicle_name,
                v.get_vehicle_type_display(),
                diag["mileage"],
                diag["year"],
                diag["age"],
                diag["service_count"],
                diag["days_since_last_service"],
                f"{diag['failure_probability']}%",
                diag["risk_level"],
                diag["prediction"],
                diag["recommended_action"],
                diag["estimated_cost"]
            ])

        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="ai_predictive_maintenance_audit.xlsx"'
        workbook.save(response)
        return response